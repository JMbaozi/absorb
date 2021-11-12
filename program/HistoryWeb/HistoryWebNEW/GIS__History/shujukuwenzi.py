from openpyxl import load_workbook
import sqlite3


def init_db(dbpath):
    sql = '''
        create table wenzifayuan
        (
        id INTEGER PRIMARY  KEY autoincrement,
        遗址名称 text,
        描述 text,
        link text ,
        遗址出土年代 text ,
        遗址地点 text ,
        遗址坐标 text,
        遗址类型 text

        ) '''  # 创建表语句
    conn = sqlite3.connect(dbpath)
    cursor = conn.cursor()
    cursor.execute(sql)
    conn.commit()
    conn.close()


def saveDataSqlite(datalist, savepath):
    conn = sqlite3.connect(savepath)
    cur = conn.cursor()
    for index in range(len(datalist)):
        datalist[index] = '"' + str(datalist[index]) + '"'
    sql = '''  insert into wenzifayuan(
                    遗址名称,描述,link,遗址出土年代,遗址地点,遗址坐标,遗址类型)
                    VALUES (%s)''' % ",".join(datalist)
    print(sql)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()


def run():
    workbook = load_workbook('东夷文字发源.xlsx')
    sheet = workbook['文字发源地']
    data = []
    path = "HISTORY_DATABASE"
    for j in range(2, 14):
        for i in range(1, 8):
            data1 = sheet.cell(j, i).value
            data.append(data1)
        saveDataSqlite(data, path)
        data = []


if __name__ == '__main__':
    run()
    # init_db("HISTORY_DATABASE")