from openpyxl import load_workbook
import sqlite3

def init_db(dbpath):
    sql = '''
        create table trace
        (
        id INTEGER PRIMARY  KEY autoincrement,
        name text,
        time text,
        Biao_place text,
        C_place text ,
        info text ,
        lon float ,
        lat float

        ) '''  # 创建表语句
    conn = sqlite3.connect(dbpath)
    cursor = conn.cursor()
    cursor.execute(sql)
    conn.commit()
    conn.close()


def saveDataSqlite(datalist, savepath):
    conn = sqlite3.connect(savepath)
    cur = conn.cursor()
    for index in range(len(datalist)):
        datalist[index] = '"' + str(datalist[index]) + '"'
    sql = '''  insert into trace(
                    name,time,Biao_place,C_place,info,lon,lat)
                    VALUES (%s)''' % ",".join(datalist)
    print(sql)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()


def run():
    workbook = load_workbook('人物轨迹.xlsx')
    sheet = workbook['Sheet1']
    data = []
    path = "HISTORY_DATABASE"
    for j in range(2, 10):
        for i in range(1, 8):
            data1 = sheet.cell(j, i).value
            data.append(data1)
        saveDataSqlite(data, path)
        data = []


if __name__ == '__main__':
    run()
    # init_db("HISTORY_DATABASE")