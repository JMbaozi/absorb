from openpyxl import load_workbook
import sqlite3

def init_db(dbpath):
    sql='''
        create table people
        (
        id INTEGER PRIMARY  KEY autoincrement,
        朝代 text,
        地区 text,
        户数 INT ,
        口数 INT ,
        人每户 FLOAT ,
        备注 text  
        ) ''' #创建表语句
    conn=sqlite3.connect(dbpath)
    cursor=conn.cursor()
    cursor.execute(sql)
    conn.commit()
    conn.close()

def saveDataSqlite(datalist,savepath):
    conn=sqlite3.connect(savepath)
    cur=conn.cursor()

    for index in range(len(datalist)):
        datalist[index] = '"' + str(datalist[index]) + '"'
    print(datalist)
    sql='''  insert into people(
                    朝代,地区,户数,口数,人每户,备注)
                    VALUES (%s)'''%",".join(datalist)
    print(sql)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()
def run():
    workbook = load_workbook('表.xlsx')
    sheet = workbook['人口']
    data=[]
    path="HISTORY_数据库"
    init_db(path)
    for j in range(2, 148):
        for i in range(1,7):
            data1 = sheet.cell(j, i).value
            if isinstance(data1, str):
                data1.strip()
                data.append(data1.strip())
            else :
                data.append(str(data1))
        saveDataSqlite(data,path)
        data = []

if __name__ == '__main__':

    run()