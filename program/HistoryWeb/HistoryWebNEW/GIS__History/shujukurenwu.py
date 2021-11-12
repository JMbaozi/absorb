from openpyxl import load_workbook
import sqlite3


def init_db(dbpath):
    sql = '''
        create table renwu
        (
        id INTEGER PRIMARY  KEY autoincrement,
        name text,
        info text,
        地点 text ,
        朝代 text ,
        人物属性 text ,
        经度 float,
        维度 float
        ) '''  # 创建表语句
    conn = sqlite3.connect(dbpath)
    cursor = conn.cursor()
    cursor.execute(sql)
    conn.commit()
    conn.close()


def saveDataSqlite(datalist, savepath):
    conn = sqlite3.connect(savepath)
    cur = conn.cursor()
    for index in range(len(datalist)):
        datalist[index] = '"' + str(datalist[index]) + '"'
    sql = '''  insert into renwu(
                    name,info,地点,朝代,人物属性,经度,维度)
                    VALUES (%s)''' % ",".join(datalist)
    print(sql)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()


def run():
    workbook = load_workbook('人物2.xlsx')
    sheet = workbook['Sheet1']
    data = []
    path = "HISTORY_DATABASE"
    for j in range(1, 273):
        for i in range(1, 8):
            data1 = sheet.cell(j, i).value
            data.append(data1)
        saveDataSqlite(data, path)
        data = []


if __name__ == '__main__':
    run()
    # init_db("HISTORY_DATABASE")