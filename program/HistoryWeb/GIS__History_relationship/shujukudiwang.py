from openpyxl import load_workbook
import sqlite3


def init_db(dbpath):
    sql='''
        create table tour
        (
        id INTEGER PRIMARY  KEY autoincrement,
        朝代 text,
        朝时间 text,
        游历时间 text ,
        参与人物 text ,
        介绍 text ,
        历经地点 text,
        坐标 text,
        地点介绍 text
          
        ) ''' #创建表语句
    conn=sqlite3.connect(dbpath)
    cursor=conn.cursor()
    cursor.execute(sql)
    conn.commit()
    conn.close()


def saveDataSqlite(datalist,savepath):
    conn=sqlite3.connect(savepath)
    cur=conn.cursor()
    for index in range(len(datalist)):
        datalist[index] = '"' + str(datalist[index]) + '"'
    sql='''  insert into tour(
                    朝代,朝时间,游历时间,参与人物,介绍,历经地点,坐标,地点介绍)
                    VALUES (%s)'''%",".join(datalist)
    print(sql)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()
def run():
    workbook = load_workbook('表1.xlsx')
    sheet = workbook['帝王巡视']
    data=[]
    path="HISTORY_DATABASE"
    for j in range(2, 32):
        for i in range(1,9):
            data1 = sheet.cell(j, i).value
            data.append(data1)
        saveDataSqlite(data,path)
        data = []
if __name__ == '__main__':
    run()
    # init_db("HISTORY_DATABASE")